import openpyxl

class HaushaltsplanExcelGenerator:
    def generate_excel(self, haushaltsplan, file_name):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet['A1'] = 'Projekt Name'
        worksheet['B1'] = 'Einnahmen'
        worksheet['C1'] = 'Ausgaben'

        row_index = 2

        for haushaltsposten in haushaltsplan.haushaltsposten:
            worksheet.cell(row=row_index, column=1, value=haushaltsposten.name)

            for projekt in haushaltsposten.projekte:
                worksheet.cell(row=row_index, column=2, value=projekt.einnahmen)
                worksheet.cell(row=row_index, column=3, value=projekt.ausgaben)
                row_index += 1

        workbook.save(file_name)
